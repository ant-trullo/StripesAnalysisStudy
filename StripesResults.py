"""This function write a xlsx file with the results of
the strip's analysis.
"""

import datetime
import numpy as np
import xlsxwriter
from openpyxl import load_workbook
from skimage.measure import regionprops_table
# import pyqtgraph as pg


class StripesResultsSpots:
    """Onlly class, does all the job"""
    def __init__(self, pseudo_cells, new_cells, roi, post_proc_folder, flag_a_b, software_version):

        wwbb      =  load_workbook(post_proc_folder + '/PseudoCytoPlasmSpots_info' + flag_a_b + '.xlsx')
        sheet_ts  =  wwbb[wwbb.sheetnames[wwbb.sheetnames.index("TS")]]

        org_list  =  []
        idxs_new  =  np.unique(new_cells[new_cells != 0])

        for k in idxs_new:
            bff  =  (new_cells == k) * pseudo_cells
            org_list.append(np.unique(bff[bff != 0]))

        workbook  =  xlsxwriter.Workbook(post_proc_folder + '/StripsStudy.xlsx')
        sheet1    =  workbook.add_worksheet("Strips")
        sheet2    =  workbook.add_worksheet("Regions")

        # wwbb  =  load_workbook(post_proc_folder + '/PseudoCytoPlasmSpots_info' + flag_a_b + '.xlsx')
        if wwbb.sheetnames[2] == "cl-Summary":
            sheet_tags  =  wwbb[wwbb.sheetnames[4]]
            list_tags   =  []
            [list_tags.append(sheet_tags.cell(row=x, column=1).value) for x in range(2, sheet_tags.max_row + 1)]
        elif wwbb.sheetnames[2] == "TS":
            sheet_tags  =  wwbb[wwbb.sheetnames[1]]
            list_tags   =  []
            [list_tags.append(sheet_tags.cell(row=x, column=1).value) for x in range(2, sheet_tags.max_row + 1)]

        sheet1.write(0, 0, "Stripe Num")
        sheet1.write(0, 1, "Tot eq sm")
        sheet1.write(0, 2, "TS Ints-1")
        sheet1.write(0, 3, "TS Ints-2")

        row_idx  =  3
        spr_lst  =  []
        for ff in range(len(org_list)):
            bff_lst  =  []
            sheet1.write(row_idx, 0, "Stripe " + str(idxs_new[ff]))
            row_idx  +=  1
            bff_lst.append(idxs_new[ff])
            for mm in org_list[ff]:
                bff_lst.append([mm, sheet_tags.cell(row=list_tags.index("Cyto_" + str(mm)) + 2, column=6).value])
                sheet1.write(row_idx, 0, "Cyto_" + str(mm))
                sheet1.write(row_idx, 1, sheet_tags.cell(row=list_tags.index("Cyto_" + str(mm)) + 2, column=6).value)
                sheet1.write(row_idx, 2, sheet_ts.cell(row=list_tags.index("Cyto_" + str(mm)) + 2, column=3).value)
                sheet1.write(row_idx, 3, sheet_ts.cell(row=list_tags.index("Cyto_" + str(mm)) + 2, column=15).value)
                row_idx  +=  1
            spr_lst.append(bff_lst)
            row_idx  += 2

        sheet1.write(0, 7, "Analysis Folder")
        sheet1.write(0, 8, post_proc_folder)

        sheet1.write(2, 7, "Software Version")
        sheet1.write(2, 8, software_version)

        sheet1.write(3, 7, "Date")
        sheet1.write(3, 8, datetime.date.today().strftime("%d%b%Y"))

        brdr_regions  =  roi.getRegion()
        rgp_nwclls    =  regionprops_table(new_cells, properties=["label", "centroid", "coords"])             # regionprops of the pseudo cells
        if rgp_nwclls["coords"][0][:, 1].min() - rgp_nwclls["coords"][0][:, 1].max() < rgp_nwclls["coords"][0][:, 2].min() - rgp_nwclls["coords"][0][:, 2].max(): # check if strips are hor or vert

            lbls_ctrs   =  np.zeros((2, len(rgp_nwclls["label"])))                                             # 3D coordinates of the centroids
            for cntr, k in enumerate(rgp_nwclls["label"]):
                lbls_ctrs[:, cntr]  =  k, rgp_nwclls["centroid-2"][cntr]                                       # matrix with labels and centroid y-coordinates

            splt_ref    =  brdr_regions[(np.abs(np.asarray(brdr_regions) - pseudo_cells.shape[1])).argmin()]
            strp1_lbls  =  lbls_ctrs[0, np.where(lbls_ctrs[1, :] < splt_ref)[0]]

        else:

            lbls_ctrs   =  np.zeros((2, len(rgp_nwclls["label"])))                                             # 3D coordinates of the centroids
            for cntr, k in enumerate(rgp_nwclls["label"]):
                lbls_ctrs[:, cntr]  =  k, rgp_nwclls["centroid-1"][cntr]                                       # matrix with labels and centroid x-coordinates

            splt_ref    =  brdr_regions[(np.abs(np.asarray(brdr_regions) - pseudo_cells.shape[2])).argmin()]
            strp1_lbls  =  lbls_ctrs[0, np.where(lbls_ctrs[1, :] < splt_ref)[0]]

        strp2_lbls  =  list(lbls_ctrs[0, :])
        for hh in strp1_lbls:
            strp2_lbls.remove(hh)

        sheet2.write(0, 0, "Region1")
        sheet2.write(2, 0, "Cyto_id")
        sheet2.write(2, 1, "Tot eq sm")
        row_idx2  =  3
        for str_id1 in strp1_lbls:
            ii = -1
            while spr_lst[ii][0] != str_id1:
                ii += 1
            for jj in spr_lst[ii][1:]:
                sheet2.write(row_idx2, 0, "Cyto_" + str(jj[0]))
                sheet2.write(row_idx2, 1, jj[1])
                row_idx2  +=  1

        sheet2.write(0, 5, "Region2")
        sheet2.write(2, 5, "Cyto_id")
        sheet2.write(2, 6, "Tot eq sm")

        row_idx3  =  3
        for str_id2 in strp2_lbls:
            ll = -1
            while spr_lst[ll][0] != str_id2:
                ll  +=  1
            for vv in spr_lst[ll][1:]:
                sheet2.write(row_idx3, 5, "Cyto_" + str(vv[0]))
                sheet2.write(row_idx3, 6, vv[1])
                row_idx3  +=  1

        workbook.close()

        self.org_list  =  org_list


class StripesResultsNucs:
    """Only class, does all the job"""
    def __init__(self, pseudo_cells, new_cells, quantif_mtx, nucs_piled, roi, post_proc_folder, raw_data_fname, ref_quantif_chs, dapi_ellipsoid_flag, software_version):

        org_list  =  []                                                                                                 # initialize list (it will be a list of lists)
        idxs_new  =  np.unique(new_cells[new_cells != 0])                                                               # tags of all the nuclei

        for k in idxs_new:
            bff  =  (new_cells == k) * pseudo_cells                                                                     # associate to each stripe the list of the tags of all the nuclei inside the stripe
            org_list.append(np.unique(bff[bff != 0]))                                                                   # add the list to the list of lists
        rgp_nucs   =  regionprops_table(nucs_piled, quantif_mtx, properties=["label", "intensity_image", "area"])       # label intensity image and area of each nucleus to estimate the average intensity
        reference  =  max(roi.getRegion()[0], roi.getRegion()[1])                                                       # roi gives the position of the 2 horizontal lines of the total roi: only the one on the bottom is important

        max_length  =  0                                                                                                # important value for excell writing
        for uu in org_list:
            max_length  =  max(max_length, len(uu))                                                                     # search the maximum number of nuclei in the stripe
        max_length  +=  2                                                                                               # add 2, just for excell visualization purpose

        rgp_new_cells           =  regionprops_table(new_cells, properties=["label", "centroid"])                       # regionprops: label and centroid to give organized tags to the stripes
        new_cells_lb_ctr        =  np.zeros((2, len(rgp_new_cells["label"])))                                           # store label and y centroid coordinate (stripes are horizontal)
        new_cells_lb_ctr[0, :]  =  rgp_new_cells["label"]                                                               # add labels
        new_cells_lb_ctr[1, :]  =  rgp_new_cells["centroid-2"]                                                          # add centroid y coordinates in corrispondent positions

        ords                                          =  np.zeros(new_cells_lb_ctr.shape[1])                                        # ordering of the strips with respect to the reference
        ords[new_cells_lb_ctr[1, :] - reference > 0]  =  np.arange(ords[new_cells_lb_ctr[1, :] - reference > 0].size, 0, -1)        # all the stripes with centroids higher than reference are positive integer nubers increasing with the distance
        ords[new_cells_lb_ctr[1, :] - reference < 0]  =  - np.arange(1, ords[new_cells_lb_ctr[1, :] - reference < 0].size + 1)      # all the stripes with centroids smaller than reference are negative integer nubers decreasing with the distance

        tot_list1  =  [["" for i1 in range(max_length + 1)] for b1 in range(len(rgp_new_cells["label"]))]               # initialize the 3 lists of list used to write the excell file
        tot_list2  =  [["" for i2 in range(max_length + 1)] for b2 in range(len(rgp_new_cells["label"]))]
        tot_list3  =  [["" for i3 in range(max_length + 1)] for b3 in range(len(rgp_new_cells["label"]))]

        for cnt_o, uff in enumerate(ords):                                                                              # put the order number in list 3
            tot_list3[cnt_o][0]  =  uff

        row_idx  =  0
        for pp in range(len(org_list)):
            tot_list1[row_idx][0]  =  "Stripe_" + str(idxs_new[pp])                                                     # add the text of the first column in the sheet in sheet 1 and 2
            tot_list2[row_idx][0]  =  "Stripe_" + str(idxs_new[pp])

            ints_stripe  =  []
            for cnt, nucs_idx in enumerate(org_list[pp]):
                tot_list1[row_idx][cnt + 1]  =  "Nuc_" + str(nucs_idx)                                                  # sheet1: add the tags of the nuclei involved in the stripe
                ll                           =  np.where(rgp_nucs["label"] == nucs_idx)[0]                              # search the coordinate of the corresponding nucleus in the regionprop table
                if ll.size != 0:
                    ll                           =  ll[0]
                    tot_list2[row_idx][cnt + 1]  =  np.sum(rgp_nucs["intensity_image"][ll]) / rgp_nucs["area"][ll]          # add average intensity in both sheet2 and 3
                    tot_list3[row_idx][cnt + 1]  =  np.sum(rgp_nucs["intensity_image"][ll]) / rgp_nucs["area"][ll]
                    ints_stripe.append(np.sum(rgp_nucs["intensity_image"][ll]) / rgp_nucs["area"][ll])                      # add into a buffer list to calculate the average of the averages in the stripe
            tot_list2[row_idx][max_length]  =  np.asarray(ints_stripe).mean()                                           # add the average of the averages in sheet2 and 3
            tot_list3[row_idx][max_length]  =  np.asarray(ints_stripe).mean()
            row_idx  +=  1

        tot_list1  =  tot_list1[::-1]                                                                                   # invert lists in order to have the text of the excell oriented as the stripes in the image
        tot_list2  =  tot_list2[::-1]
        tot_list3  =  tot_list3[::-1]

        workbook    =  xlsxwriter.Workbook(post_proc_folder + '/StripsStudyNucs.xlsx')
        sheet1      =  workbook.add_worksheet("Strips Nucs")
        sheet2      =  workbook.add_worksheet("Strips Ints")
        sheet3      =  workbook.add_worksheet("Strips Ordered")
        sheet_info  =  workbook.add_worksheet("Info")

        sheet_info.write(0, 0, "Reference Position")
        sheet_info.write(0, 1, reference)

        sheet_info.write(2, 0, "Reference Channel")
        sheet_info.write(3, 0, "Quantitation Channel")

        if ref_quantif_chs[0] == 0:
            sheet_info.write(2, 1, "Spots A")
        elif ref_quantif_chs[0] == 1:
            sheet_info.write(2, 1, "Spots B")
        elif ref_quantif_chs[0] == 2:
            sheet_info.write(2, 1, "None")

        if ref_quantif_chs[1] == 0:
            sheet_info.write(3, 1, "Spots A")
        elif ref_quantif_chs[1] == 1:
            sheet_info.write(3, 1, "Spots B")

        sheet_info.write(4, 0, "Nuclei")
        sheet_info.write(4, 1, dapi_ellipsoid_flag)

        sheet_info.write(6, 0, "File")
        sheet_info.write(6, 1, raw_data_fname)

        sheet_info.write(8, 0, "date")
        sheet_info.write(8, 1, datetime.date.today().strftime("%d%b%Y"))

        sheet_info.write(10, 0, "Software Version")
        sheet_info.write(10, 1, software_version)

        sheet1.write(0, 0, "Stripe Tag")
        sheet2.write(0, 0, "Stripe Tag")
        sheet2.write(0, max_length, "Average in Stripe")
        sheet3.write(0, 0, "Stripe Order")
        sheet3.write(0, max_length, "Average in Stripe")

        for cnt1 in range(len(tot_list1)):
            for cnt1_s in range(len(tot_list1[cnt1])):
                if tot_list1[cnt1][cnt1_s] != "":                                                                       # some element in the lists should be empty to respect the geometry of the excell file: empty codded by "" in this implementation
                    sheet1.write(1 + cnt1, cnt1_s, tot_list1[cnt1][cnt1_s])

        for cnt2 in range(len(tot_list2)):
            for cnt2_s in range(len(tot_list2[cnt2])):
                if tot_list2[cnt2][cnt2_s] != "":
                    sheet2.write(1 + cnt2, cnt2_s, tot_list2[cnt2][cnt2_s])

        for cnt3 in range(len(tot_list3)):
            for cnt3_s in range(len(tot_list3[cnt3])):
                if tot_list3[cnt3][cnt3_s] != "":
                    sheet3.write(1 + cnt3, cnt3_s, tot_list3[cnt3][cnt3_s])

        workbook.close()
