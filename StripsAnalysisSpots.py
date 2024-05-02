"""This widgets serves as a standalone postprocessing tool
for results of smFish_Software (https://github.com/ant-trullo/smFiSH_software).

It reorganizes results in strips.
author: antonio.trullo@igmm.cnrs.fr
"""

import sys
import os
import numpy as np
from openpyxl import load_workbook
from PyQt5.QtCore import Qt
from PyQt5 import QtWidgets
import pyqtgraph as pg

import RawDataLoader2
import StripesResults
import PseudoCellsOrganize
import TagsAndPositions


class StripsAnalysisSpots(QtWidgets.QWidget):
    """Popup tool to study activation in strips."""
    # def __init__(self, raw_data_fname, post_proc_folder, flag_tssl_mbm, flag_hor_vert, flag_a_b, flag_a_b4visual):
    def __init__(self):
        QtWidgets.QWidget.__init__(self)

        post_proc_folder  =  '/home/atrullo/Dropbox/StripesAnalysisStudy/Analysis center'             # path of the folder with the analysis
        raw_data_fname    =  '/home/atrullo/Dropbox/StripesAnalysisStudy/Analysis center/LM107_07052022_snaShaE_E1.tif'             # path of the raw data file
        flag_hor_vert     =  "hor"          # or "vert"  depending if you want horizontal or vertical strips
        flag_tssl_mbm     =  "tassels"      # membrane or "tassels" depending if you want to study on a meembrane analysis or a tassels (voronoi pseudo cells) analysis
        flag_a_b          =  "_a"           # if you want to study spots in the a channel of the analys or "_b" if in the b channel oof the analysis
        flag_a_b4visual   =  flag_a_b

        wb_cyto   =  load_workbook(post_proc_folder + '/PseudoCytoPlasmSpots_info' + flag_a_b + '.xlsx')
        wb_cyto1  =  wb_cyto[wb_cyto.sheetnames[0]]

        if flag_a_b4visual == "_a":
            raw_spts   =  RawDataLoader2.RawDataLoader(raw_data_fname, np.load(post_proc_folder + '/chs_spts_nucs.npy')).spts_a[wb_cyto1["I1"].value:wb_cyto1["I3"].value]
        elif flag_a_b4visual == "_b":
            raw_spts   =  RawDataLoader2.RawDataLoader(raw_data_fname, np.load(post_proc_folder + '/chs_spts_nucs.npy')).spts_b[wb_cyto1["I1"].value:wb_cyto1["I3"].value]
        if os.path.isfile(post_proc_folder + '/roi_crop.npy'):
            roi_crop  =  np.load(post_proc_folder + '/roi_crop.npy')
            raw_spts  =  raw_spts[:, roi_crop[0]:roi_crop[2], roi_crop[1]:roi_crop[3]]

        if flag_tssl_mbm == "tassels":
            pseudo_cells  =  np.load(post_proc_folder + '/pseudo_cells.npy')
        elif flag_tssl_mbm == "membrane":
            pseudo_cells  =  np.load(post_proc_folder + '/pseudo_memb_cells.npy')

        if flag_hor_vert == "hor":
            new_cells  =  PseudoCellsOrganize.CellsStripesHor(pseudo_cells).new_cells
        elif flag_hor_vert == "vert":
            new_cells  =  PseudoCellsOrganize.CellsStripesVer(pseudo_cells).new_cells

        tag_poss  =  TagsAndPositions.TagsAndPositions(new_cells)

        ksf_h  =  np.load('keys_size_factor.npy')[0]
        ksf_w  =  np.load('keys_size_factor.npy')[1]

        mycmap      =  np.fromfile("mycmap.bin", "uint16").reshape((10000, 3))      # / 255.0
        colors4map  =  []
        for k in range(mycmap.shape[0]):
            colors4map.append(mycmap[k, :])
        colors4map[0]  =  np.array([0, 0, 0])
        rnd_cmap       =  pg.ColorMap(np.linspace(0, 1, pseudo_cells.max()), color=colors4map)

        fname_lbl  =  QtWidgets.QLabel("File: " + post_proc_folder)
        fname_lbl.setToolTip("Name of the file you are working on")

        tabs              =  QtWidgets.QTabWidget()
        tab_pseudo_cells  =  QtWidgets.QWidget()
        tab_spts_raw      =  QtWidgets.QWidget()

        frame_new_cell  =  pg.ImageView(self, name="FramePseudoCells")
        frame_new_cell.ui.roiBtn.hide()
        frame_new_cell.ui.menuBtn.hide()
        frame_new_cell.timeLine.sigPositionChanged.connect(self.update_frames_from_new)
        frame_new_cell.setImage(new_cells)
        frame_new_cell.setColorMap(rnd_cmap)

        tags4stripes  =  list()
        for mm in range(tag_poss.tags.size):
            tags4stripes.append(pg.TextItem(str(int(tag_poss.tags[mm])), color=[255, 0, 0]))

        for gg in range(tag_poss.tags.size):
            frame_new_cell.addItem(tags4stripes[gg])
            tags4stripes[gg].setPos(tag_poss.poss[gg, 0, 0], tag_poss.poss[gg, 1, 0])

        frame_spts_raw  =  pg.ImageView(self)
        frame_spts_raw.view.setXLink("FramePseudoCells")
        frame_spts_raw.view.setYLink("FramePseudoCells")
        frame_spts_raw.timeLine.sigPositionChanged.connect(self.update_frames_from_spts)
        frame_spts_raw.setImage(raw_spts)

        frame_spts_raw_box  =  QtWidgets.QHBoxLayout()
        frame_spts_raw_box.addWidget(frame_spts_raw)

        frame_new_cells_box  =  QtWidgets.QHBoxLayout()
        frame_new_cells_box.addWidget(frame_new_cell)

        tab_pseudo_cells.setLayout(frame_new_cells_box)
        tab_spts_raw.setLayout(frame_spts_raw_box)

        tabs.addTab(tab_pseudo_cells, "Pseudo Cells")
        tabs.addTab(tab_spts_raw, "Raw Data")

        if flag_hor_vert == "hor":
            roi_new  =  pg.LinearRegionItem(orientation=True)
            roi_raw  =  pg.LinearRegionItem(orientation=True)
            roi_new.sigRegionChanged.connect(self.update_roiraw_from_new)
            roi_raw.sigRegionChanged.connect(self.update_roinew_from_raw)
            # roi.setY(brd_coord.y_pos)
            frame_new_cell.addItem(roi_new)
            frame_spts_raw.addItem(roi_raw)
        elif flag_hor_vert == "vert":
            roi_new  =  pg.LinearRegionItem(orientation=False)
            roi_raw  =  pg.LinearRegionItem(orientation=False)
            roi_new.sigRegionChanged.connect(self.update_roiraw_from_new)
            roi_raw.sigRegionChanged.connect(self.update_roinew_from_raw)
            # roi.setX(brd_coord.x_pos)
            frame_new_cell.addItem(roi_new)
            frame_spts_raw.addItem(roi_raw)

        save_stripes_btn  =  QtWidgets.QPushButton("Save", self)
        save_stripes_btn.setToolTip("Save stripes info")
        save_stripes_btn.setFixedSize(int(ksf_h * 60), int(ksf_w * 25))
        save_stripes_btn.clicked.connect(self.save_stripes)

        commands_box  =  QtWidgets.QHBoxLayout()
        commands_box.addStretch()
        commands_box.addWidget(save_stripes_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(tabs)
        layout.addLayout(commands_box)

        self.frame_new_cell    =  frame_new_cell
        self.frame_spts_raw    =  frame_spts_raw
        self.post_proc_folder  =  post_proc_folder
        self.pseudo_cells      =  pseudo_cells
        self.new_cells         =  new_cells
        self.roi_new           =  roi_new
        self.roi_raw           =  roi_raw
        self.flag_hor_vert     =  flag_hor_vert
        self.software_version  =  "smiFish_Software_3.1"
        self.flag_a_b          =  flag_a_b
        self.tag_poss          =  tag_poss
        self.tags4stripes      =  tags4stripes

        self.setLayout(layout)
        self.setGeometry(300, 300, int(ksf_h * 450), int(ksf_w * 450))
        self.setWindowTitle("Stripes Analysis")
        self.show()

    def update_frames_from_new(self):
        """Synchronize the frame raw spts from changes on pseudocells."""
        self.frame_spts_raw.setCurrentIndex(self.frame_new_cell.currentIndex)
        for gg in range(self.tag_poss.tags.size):
            self.tags4stripes[gg].setPos(self.tag_poss.poss[gg, 0, self.frame_new_cell.currentIndex], self.tag_poss.poss[gg, 1, self.frame_new_cell.currentIndex])

    def update_frames_from_spts(self):
        """Synchronize the frame pseudocells from changes on raw spts."""
        self.frame_new_cell.setCurrentIndex(self.frame_spts_raw.currentIndex)

    def update_roinew_from_raw(self):
        """Keep the two roi updated from each other."""
        self.roi_new.setRegion(self.roi_raw.getRegion())

    def update_roiraw_from_new(self):
        """Keep the two roi updated from each other."""
        self.roi_raw.setRegion(self.roi_new.getRegion())

    def save_stripes(self):
        """Save stripes info in xlsx file."""
        StripesResults.StripesResultsSpots(self.pseudo_cells, self.new_cells, self.roi_new, self.post_proc_folder, self.flag_a_b, self.software_version)


class BinaryChoice(QtWidgets.QDialog):
    """Choose the spots channel to analyse."""
    def __init__(self, text_list, parent=None):
        super().__init__(parent)

        text_label  =  text_list[0]
        text_btn1   =  text_list[1]
        text_btn2   =  text_list[2]
        text_tip1   =  text_list[3]
        text_tip2   =  text_list[4]
        text_title  =  text_list[5]

        ksf_h  =  np.load('keys_size_factor.npy')[0]
        ksf_w  =  np.load('keys_size_factor.npy')[1]

        choose_lbl  =  QtWidgets.QLabel(text_label, self)
        choose_lbl.setFixedSize(int(ksf_h * 180), int(ksf_w * 22))

        coeff_length  =  5 * max(len(text_tip1), len(text_tip2))

        choose_first_btn  =  QtWidgets.QPushButton(text_btn1, self)
        choose_first_btn.setToolTip(text_tip1)
        choose_first_btn.setFixedSize(int(ksf_h * coeff_length), int(ksf_w * 25))
        choose_first_btn.clicked.connect(self.choose_first)

        choose_second_btn  =  QtWidgets.QPushButton(text_btn2, self)
        choose_second_btn.setToolTip(text_tip2)
        choose_second_btn.setFixedSize(int(ksf_h * coeff_length), int(ksf_w * 25))
        choose_second_btn.clicked.connect(self.choose_second)

        choose_box  =  QtWidgets.QHBoxLayout()
        choose_box.addWidget(choose_first_btn)
        choose_box.addStretch()
        choose_box.addWidget(choose_second_btn)

        layout  =  QtWidgets.QVBoxLayout()
        layout.addWidget(choose_lbl)
        layout.addLayout(choose_box)

        self.setWindowModality(Qt.ApplicationModal)
        self.setLayout(layout)
        self.setGeometry(300, 300, int(ksf_h * 200), int(ksf_w * 25))
        self.setWindowTitle(text_title)

    def choose_first(self):
        """Choose flag first."""
        self.flag_first_second  =  0
        self.close()

    def choose_second(self):
        """Choose flag second."""
        self.flag_first_second  =  1
        self.close()

    def params(self):
        """Function to send choice."""
        return self.flag_first_second

    @staticmethod
    def getFlag(parent=None):
        """Send choice."""
        dialog  =  BinaryChoice(parent)
        result  =  dialog.exec_()
        flag    =  dialog.params()
        return flag


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ex = StripsAnalysisSpots()
    sys.exit(app.exec_())